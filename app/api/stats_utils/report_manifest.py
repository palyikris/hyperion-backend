"""
Geospatial Cleanup Manifest Report Generator

This module generates Excel reports containing detailed information about media
and their associated trash detections. Each row represents a single detection
with full geospatial, AI, forensic, and audit metadata.

The report is designed for cleanup verification, environmental audits, and
operational planning.
"""

from io import BytesIO
from datetime import datetime, timezone, timedelta
from typing import BinaryIO

import pandas as pd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.models.db.Media import Media
from app.models.db.Detection import Detection
from app.models.upload.MediaStatus import MediaStatus


def _extract_technical_metadata(tech_meta: dict | None) -> tuple[str, str, str]:
    """
    Extract camera make, model, and date_taken from technical_metadata JSON.
    
    Args:
        tech_meta: The technical_metadata dict from Media model
        
    Returns:
        Tuple of (make, model, date_taken) with "N/A" as default
    """
    if not tech_meta:
        return "N/A", "N/A", "N/A"
    
    make = tech_meta.get("make", "N/A")
    model = tech_meta.get("model", "N/A")
    date_taken = tech_meta.get("date_taken", "N/A")
    
    return make, model, date_taken


def _build_hf_url(hf_path: str | None, uploader_id: str, media_id: str) -> str:
    """
    Construct the full Hugging Face URL for a media file.
    
    Args:
        hf_path: The hf_path from Media model (may be None)
        uploader_id: The user ID who uploaded the media
        media_id: The media UUID
        
    Returns:
        Full URL to the image on Hugging Face, or "N/A" if hf_path is missing
    """
    if not hf_path:
        return "N/A"

    base_url = "https://huggingface.co/datasets/palyikris/hyperion-media/resolve/main"
    return f"{base_url}/{hf_path}"


async def generate_manifest_data(
    db: AsyncSession, user_id: str, days: int = 30
) -> list[dict]:
    """
    Generate flattened data for the Geospatial Cleanup Manifest.
    
    This function queries all READY media for the authenticated user within
    the specified time window, and flattens each detection into a separate row.
    
    Args:
        db: Async database session
        user_id: ID of the authenticated user
        days: Number of days to look back from now (default: 30)
        
    Returns:
        List of dictionaries, each representing one detection with all metadata
        
    Data Structure:
        Each row contains:
        - Core Identification: Media ID, Detection ID, Filename
        - Geospatial: Latitude, Longitude, Altitude, Address
        - AI Insights: Object Label, Confidence (%), Area (sqm)
        - Forensics: Camera Make, Model, Date Taken
        - Audit: Assigned Titan, Image URL
    """
    # Calculate date threshold
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)

    # Query media with eager-loaded detections
    query = (
        select(Media)
        .options(selectinload(Media.detections))  # Eager load detections
        .where(Media.uploader_id == user_id)       # Security: User-scoped
        .where(Media.status == MediaStatus.READY)  # Only successfully processed
        .where(Media.created_at >= cutoff_date)    # Time filter
        .order_by(Media.created_at.desc())         # Most recent first
    )

    result = await db.execute(query)
    media_list = result.scalars().all()

    # Flatten: one row per detection
    rows = []
    for media in media_list:
        # Extract technical metadata once per media
        make, model, date_taken = _extract_technical_metadata(media.technical_metadata)

        filename = media.initial_metadata.get("filename", "N/A") if media.initial_metadata else "N/A"

        # Build image URL
        image_url = _build_hf_url(media.hf_path, media.uploader_id, str(media.id))

        # Create one row for each detection
        for detection in media.detections:
            row = {
                # Core Identification
                "Media ID": str(media.id),
                "Detection ID": str(detection.id),
                "Filename": filename,
                # Geospatial Data
                "Latitude": media.lat if media.lat is not None else "N/A",
                "Longitude": media.lng if media.lng is not None else "N/A",
                "Drone Altitude (m)": (
                    media.altitude if media.altitude is not None else "N/A"
                ),
                "Address": media.address if media.address else "N/A",
                # AI Insights
                "Object Label": detection.label,
                "Confidence (%)": round(detection.confidence * 100, 2),
                "Area (sqm)": (
                    detection.area_sqm if detection.area_sqm is not None else "N/A"
                ),
                # Forensics (EXIF Data)
                "Camera Make": make,
                "Camera Model": model,
                "Date Taken": date_taken,
                # Audit & Verification
                "Assigned Titan": (
                    media.assigned_worker if media.assigned_worker else "N/A"
                ),
                "Image URL": image_url,
                "Upload Date": (
                    media.created_at.strftime("%Y-%m-%d %H:%M:%S UTC")
                    if media.created_at
                    else "N/A"
                ),
                # Editable Fields (unlocked in protected worksheet)
                "Field Notes": "",
                "Cleanup Status": "Pending",
            }
            rows.append(row)

    return rows


import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def create_excel_file(data: list[dict]) -> BytesIO:
    """
    Advanced Excel generator with Hyperion branding, dashboard, filters, and cell protection.

    Features:
    - Dashboard worksheet with summary statistics and worker stats
    - Manifest worksheet with Hyperion Indigo styling
    - Google Maps navigation hyperlinks
    - 3-color scale conditional formatting on Confidence column
    - Sheet protection (Field Notes and Cleanup Status unlocked)
    """
    df = pd.DataFrame(data)
    buffer = BytesIO()

    # Hyperion Brand Colors
    HYPERION_BRAND = "1A5F54"
    HYPERION_WHITE = "F8F9F4"
    BORDER_COLOR = "4B5563"

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ==================== WORKSHEET 1: DASHBOARD ====================
        # Create summary statistics
        total_detections = len(df)
        unique_media = df["Media ID"].nunique()
        avg_confidence = df["Confidence (%)"].mean()
        unique_labels = df["Object Label"].nunique()
        worker_counts = df["Assigned Titan"].value_counts().to_dict()

        dashboard_data = {
            "Metric": [
                "Total Detections",
                "Unique Media Files",
                "Average Detection Confidence",
                "Unique Object Labels",
                "",
                "Worker Assignment Summary",
            ],
            "Value": [
                total_detections,
                unique_media,
                f"{avg_confidence:.2f}%",
                unique_labels,
                "",
                "",
            ],
        }

        # Add worker stats
        for idx, (worker, count) in enumerate(worker_counts.items()):
            dashboard_data["Metric"].append(f"  {worker or 'Unassigned'}")
            dashboard_data["Metric"].append("")  # Placeholder
            dashboard_data["Value"].insert(
                len(dashboard_data["Value"]), count
            )  # Insert count value

        dashboard_df = pd.DataFrame(
            {
                "Metric": [
                    "Total Detections",
                    "Unique Media Files",
                    "Average Detection Confidence",
                    "Unique Object Labels",
                    "",
                    "Worker Assignment Summary",
                ]
                + [f"  {w or 'Unassigned'}" for w in worker_counts.keys()]
            }
        )

        dashboard_df.to_excel(writer, index=False, sheet_name="Dashboard")
        dashboard_sheet = writer.sheets["Dashboard"]

        # Style the Dashboard
        dashboard_header_fill = PatternFill(
            start_color=HYPERION_BRAND, end_color=HYPERION_BRAND, fill_type="solid"
        )
        dashboard_header_font = Font(color=HYPERION_WHITE, bold=True, size=12)
        dashboard_header_alignment = Alignment(horizontal="left", vertical="center")

        thin_border = Side(border_style="thin", color=BORDER_COLOR)
        header_border = Border(
            bottom=thin_border, top=thin_border, left=thin_border, right=thin_border
        )

        # Style header row
        for col in range(1, 3):
            cell = dashboard_sheet.cell(row=1, column=col)
            cell.fill = dashboard_header_fill
            cell.font = dashboard_header_font
            cell.alignment = dashboard_header_alignment
            cell.border = header_border

        # Add metric values
        if total_detections > 0:
            dashboard_sheet["B2"] = total_detections
            dashboard_sheet["B3"] = unique_media
            dashboard_sheet["B4"] = f"{avg_confidence:.2f}%"
            dashboard_sheet["B5"] = unique_labels

        row_idx = 7
        for worker, count in worker_counts.items():
            dashboard_sheet[f"B{row_idx}"] = count
            row_idx += 1

        # Format dashboard columns
        dashboard_sheet.column_dimensions["A"].width = 35
        dashboard_sheet.column_dimensions["B"].width = 20

        # ==================== WORKSHEET 2: MANIFEST ====================
        # Insert Navigation column (Google Maps links)
        df_with_nav = df.copy()
        df_with_nav.insert(
            3,
            "Navigation",
            df_with_nav.apply(
                lambda row: (
                    f"=HYPERLINK(\"https://maps.google.com/maps?q={row['Latitude']},{row['Longitude']}\",\"Open Maps\")"
                    if row["Latitude"] != "N/A" and row["Longitude"] != "N/A"
                    else "N/A"
                ),
                axis=1,
            ),
        )

        df_with_nav.to_excel(writer, index=False, sheet_name="Cleanup Manifest")
        worksheet = writer.sheets["Cleanup Manifest"]

        # Setup table dimensions
        max_row = len(df_with_nav) + 1
        max_col = len(df_with_nav.columns)

        # Hyperion Indigo header styling
        header_fill = PatternFill(
            start_color=HYPERION_BRAND, end_color=HYPERION_BRAND, fill_type="solid"
        )
        header_font = Font(color=HYPERION_WHITE, bold=True, size=11)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        thin_border = Side(border_style="thin", color=BORDER_COLOR)
        header_border = Border(
            bottom=thin_border, top=thin_border, left=thin_border, right=thin_border
        )

        # Apply header styling
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        # Get indices of unlocked columns
        unlocked_columns = []
        for idx, col_name in enumerate(df_with_nav.columns, 1):
            if col_name in ["Field Notes", "Cleanup Status"]:
                unlocked_columns.append(idx)

        # Apply number format to Confidence column and add conditional formatting
        confidence_col_idx = None
        for idx, col_name in enumerate(df_with_nav.columns, 1):
            if col_name == "Confidence (%)":
                confidence_col_idx = idx
                column_letter = get_column_letter(idx)

                # Apply 3-color scale (Red to Yellow to Green)
                color_scale = ColorScaleRule(
                    start_type="min",
                    start_color="FF0000",  # Red
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFFF00",  # Yellow
                    end_type="max",
                    end_color="00B050",  # Green
                )
                worksheet.conditional_formatting.add(
                    f"{column_letter}2:{column_letter}{max_row}",
                    color_scale,
                )
                break

        # Apply formatting and protection to all cells
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Light borders for readability
                cell.border = Border(
                    left=thin_border, right=thin_border, bottom=thin_border
                )

                # Make hyperlinks blue
                column_name = str(worksheet.cell(row=1, column=col_idx).value)
                if ("URL" in column_name or "Navigation" in column_name) and cell.value:
                    if isinstance(cell.value, str) and cell.value.startswith(
                        "=HYPERLINK"
                    ):
                        cell.font = Font(color="0000FF", underline="single")
                    else:
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0000FF", underline="single")

                # Row zebra-striping
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
                    )

                # Set protection: unlock only Field Notes and Cleanup Status
                if col_idx in unlocked_columns:
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)

        # Also unprotect headers for unlocked columns
        for col_idx in unlocked_columns:
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.protection = Protection(locked=False)

        # Intelligent column width formatting
        for idx, column in enumerate(worksheet.columns, 1):
            column_name = str(column[0].value)
            max_length = 0
            column_letter = get_column_letter(idx)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 4, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Enable AutoFilters & Freeze Panes
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

        # Enable sheet protection with password (optional: user can specify)
        worksheet.protection.sheet = True
        worksheet.protection.password = None  # Remove or set a password
        worksheet.protection.enable()

    buffer.seek(0)
    return buffer
